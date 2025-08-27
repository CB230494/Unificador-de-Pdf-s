# =========================
# 📋 Asistencia – Público + Admin (admin oculto hasta login)
# =========================
import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import date, time
import sqlite3
from pathlib import Path

st.set_page_config(page_title="Asistencia – Registro y Admin", layout="wide")

# ---------- DB (SQLite) ----------
DB_PATH = "asistencia.db"

def get_conn():
    Path(DB_PATH).parent.mkdir(parents=True, exist_ok=True)
    return sqlite3.connect(DB_PATH, check_same_thread=False, timeout=30)

def init_db():
    with get_conn() as conn:
        conn.execute("PRAGMA journal_mode=WAL;")
        conn.execute("PRAGMA synchronous=NORMAL;")
        conn.execute("""
        CREATE TABLE IF NOT EXISTS asistencia(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            created_at TEXT NOT NULL DEFAULT (datetime('now','localtime')),
            nombre TEXT, cedula TEXT, institucion TEXT,
            cargo TEXT, telefono TEXT,
            genero TEXT, sexo TEXT, edad TEXT
        );
        """)

def insert_row(row):
    with get_conn() as conn:
        conn.execute("""INSERT INTO asistencia
            (nombre, cedula, institucion, cargo, telefono, genero, sexo, edad)
            VALUES (?,?,?,?,?,?,?,?)""",
            (row["Nombre"], row["Cédula de Identidad"], row["Institución"],
             row["Cargo"], row["Teléfono"], row["Género"], row["Sexo"], row["Rango de Edad"])
        )

def fetch_all_df(include_id=True):
    with get_conn() as conn:
        df = pd.read_sql_query("""
            SELECT id,
                   nombre  AS 'Nombre',
                   cedula  AS 'Cédula de Identidad',
                   institucion AS 'Institución',
                   cargo   AS 'Cargo',
                   telefono AS 'Teléfono',
                   genero  AS 'Género',
                   sexo    AS 'Sexo',
                   edad    AS 'Rango de Edad'
            FROM asistencia
            ORDER BY id ASC
        """, conn)
    if not df.empty:
        df.insert(0, "Nº", range(1, len(df)+1))
        if not include_id:
            df = df.drop(columns=["id"])
    else:
        cols = ["Nº","Nombre","Cédula de Identidad","Institución","Cargo","Teléfono","Género","Sexo","Rango de Edad"]
        if include_id: cols.insert(1, "id")
        df = pd.DataFrame(columns=cols)
    return df

def update_row_by_id(row_id:int, row:dict):
    with get_conn() as conn:
        conn.execute("""
            UPDATE asistencia
               SET nombre=?, cedula=?, institucion=?, cargo=?, telefono=?, genero=?, sexo=?, edad=?
             WHERE id=?""",
            (row["Nombre"], row["Cédula de Identidad"], row["Institución"],
             row["Cargo"], row["Teléfono"], row["Género"], row["Sexo"], row["Rango de Edad"], row_id)
        )

def delete_rows_by_ids(ids):
    if not ids: return
    with get_conn() as conn:
        q = ",".join("?" for _ in ids)
        conn.execute(f"DELETE FROM asistencia WHERE id IN ({q})", ids)

def delete_all_rows():
    with get_conn() as conn:
        conn.execute("DELETE FROM asistencia;")

init_db()

# ---------- Login admin en la barra lateral ----------
if "is_admin" not in st.session_state:
    st.session_state.is_admin = False

with st.sidebar:
    st.markdown("### 🔐 Acceso administrador")
    if not st.session_state.is_admin:
        pwd = st.text_input("Contraseña", type="password", placeholder="••••••••")
        if st.button("Ingresar"):
            if pwd == "Sembremos23":
                st.session_state.is_admin = True
                st.success("Acceso concedido.")
                st.rerun()
            else:
                st.error("Contraseña incorrecta.")
    else:
        st.success("Sesión de administrador activa")
        if st.button("Cerrar sesión"):
            st.session_state.is_admin = False
            st.rerun()

# ---------- Contenido PÚBLICO ----------
st.markdown("# 📋 Asistencia – Registro")
st.markdown("### ➕ Agregar")
with st.form("form_asistencia_publico", clear_on_submit=True):
    c1, c2, c3 = st.columns([1.2, 1, 1])
    nombre      = c1.text_input("Nombre")
    cedula      = c2.text_input("Cédula de Identidad")
    institucion = c3.text_input("Institución")

    c4, c5 = st.columns([1, 1])
    cargo    = c4.text_input("Cargo")
    telefono = c5.text_input("Teléfono")

    st.markdown("#### ")
    gcol, scol, ecol = st.columns([1.1, 1.5, 1.5])
    genero = gcol.radio("Género", ["F", "M", "LGBTIQ+"], horizontal=True)
    sexo   = scol.radio("Sexo (Hombre, Mujer o Intersex)", ["H", "M", "I"], horizontal=True)
    edad   = ecol.radio("Rango de Edad", ["18 a 35 años", "36 a 64 años", "65 años o más"], horizontal=True)

    submitted = st.form_submit_button("➕ Agregar", use_container_width=True)
    if submitted:
        if not nombre.strip():
            st.warning("Ingresa al menos el nombre.")
        else:
            fila = {
                "Nombre": nombre.strip(),
                "Cédula de Identidad": cedula.strip(),
                "Institución": institucion.strip(),
                "Cargo": cargo.strip(),
                "Teléfono": telefono.strip(),
                "Género": genero,
                "Sexo": sexo,
                "Rango de Edad": edad
            }
            insert_row(fila)
            st.success("Registro guardado.")

st.markdown("### 📥 Registros recibidos")
df_pub = fetch_all_df(include_id=False)
if not df_pub.empty:
    st.dataframe(
        df_pub[["Nº","Nombre","Cédula de Identidad","Institución","Cargo","Teléfono","Género","Sexo","Rango de Edad"]],
        use_container_width=True, hide_index=True
    )
else:
    st.info("Aún no hay registros guardados.")

# ---------- Contenido ADMIN ----------
if st.session_state.is_admin:
    st.markdown("---")
    st.markdown("# 🛠️ Panel del Administrador")

    # Encabezado para Excel
    st.markdown("### 🧾 Datos de encabezado (Excel)")
    col1, col2 = st.columns([1,1])
    with col1:
        fecha_evento = st.date_input("Fecha", value=date.today())
        lugar = st.text_input("Lugar", value="")
        estrategia = st.text_input("Estrategia o Programa", value="Estrategia Sembremos Seguridad")
    with col2:
        hora_inicio = st.time_input("Hora Inicio", value=time(9,0))
        hora_fin = st.time_input("Hora Finalización", value=time(12,10))
        delegacion = st.text_input("Dirección / Delegación Policial", value="")

    st.markdown("### 📝 Anotaciones y Acuerdos (para el Excel)")
    a_col, b_col = st.columns(2)
    anotaciones = a_col.text_area("Anotaciones Generales", height=220, placeholder="Escribe las anotaciones generales…")
    acuerdos    = b_col.text_area("Acuerdos", height=220, placeholder="Escribe los acuerdos…")

    st.markdown("### 👥 Registros y edición")
    df_all = fetch_all_df(include_id=True)

    if df_all.empty:
        st.info("Aún no hay registros guardados.")
    else:
        editable = df_all.copy()
        editable["Seleccionar"] = False

        edited = st.data_editor(
            editable[["Nº","Nombre","Cédula de Identidad","Institución","Cargo","Teléfono",
                      "Género","Sexo","Rango de Edad","Seleccionar"]],
            hide_index=True,
            use_container_width=True,
            column_config={
                "Nº": st.column_config.NumberColumn("Nº", disabled=True),
                "Seleccionar": st.column_config.CheckboxColumn("Seleccionar", help="Marca para eliminar"),
                "Género": st.column_config.SelectboxColumn("Género", options=["F","M","LGBTIQ+"]),
                "Sexo": st.column_config.SelectboxColumn("Sexo", options=["H","M","I"]),
                "Rango de Edad": st.column_config.SelectboxColumn("Rango de Edad",
                    options=["18 a 35 años","36 a 64 años","65 años o más"])
            },
            key="tabla_admin_editable"
        )

        c1, c2, c3, c4 = st.columns([1.2, 1.2, 1.2, 2])
        btn_save = c1.button("💾 Guardar cambios", use_container_width=True)
        btn_delete = c2.button("🗑️ Eliminar seleccionados", use_container_width=True)
        confirm_all = c4.checkbox("Confirmar vaciado total", value=False)
        btn_clear = c3.button("🧹 Vaciar todos", use_container_width=True)

        if btn_save:
            changes = 0
            for idx in edited.index:
                if idx >= len(df_all): continue
                row_orig = df_all.loc[idx]
                row_new = edited.loc[idx]
                fields = ["Nombre","Cédula de Identidad","Institución","Cargo","Teléfono","Género","Sexo","Rango de Edad"]
                if any(str(row_orig[f]) != str(row_new[f]) for f in fields):
                    update_row_by_id(int(row_orig["id"]), {f: row_new[f] for f in fields})
                    changes += 1
            st.success(f"Se guardaron {changes} cambio(s).") if changes else st.info("No hay cambios para guardar.")
            if changes: st.rerun()

        if btn_delete:
            idx_sel = edited.index[edited["Seleccionar"] == True].tolist()
            ids = df_all.iloc[idx_sel]["id"].tolist()
            if ids:
                delete_rows_by_ids(ids); st.success(f"Eliminadas {len(ids)} fila(s)."); st.rerun()
            else:
                st.info("No hay filas seleccionadas para eliminar.")

        if btn_clear:
            if confirm_all:
                delete_all_rows(); st.success("Se vaciaron todos los registros."); st.rerun()
            else:
                st.warning("Marca 'Confirmar vaciado total' para continuar.")

    # ===== Excel en UNA HOJA =====
    st.markdown("### ⬇️ Excel oficial (una sola hoja)")
    

    def build_excel_oficial_single(
        fecha: date, lugar: str, hora_ini: time, hora_fin: time,
        estrategia: str, delegacion: str, rows_df: pd.DataFrame,
        anotaciones_txt: str, acuerdos_txt: str
    ) -> bytes:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from openpyxl.drawing.image import Image as XLImage
            from openpyxl.utils import get_column_letter
            from pathlib import Path as _Path
        except Exception:
            st.error("Falta 'openpyxl' en requirements.txt")
            return b""

        # Estilos / colores
        azul_banda = "1F3B73"
        gris_head  = "D9D9D9"
        celda_fill = PatternFill("solid", fgColor=gris_head)
        banda_fill = PatternFill("solid", fgColor=azul_banda)
        th_font    = Font(bold=True)
        title_font = Font(bold=True, size=12)
        h1_font    = Font(bold=True, size=14)
        center     = Alignment(horizontal="center", vertical="center", wrap_text=True)
        right      = Alignment(horizontal="right",  vertical="center")
        left       = Alignment(horizontal="left",   vertical="center", wrap_text=True)
        thin       = Side(style="thin", color="000000")
        border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

        def outline_box(r1, c1, r2, c2):
            for c in range(c1, c2+1):
                t = ws.cell(row=r1, column=c)
                t.border = Border(top=thin, left=t.border.left, right=t.border.right, bottom=t.border.bottom)
                b = ws.cell(row=r2, column=c)
                b.border = Border(bottom=thin, left=b.border.left, right=b.border.right, top=b.border.top)
            for r in range(r1, r2+1):
                l = ws.cell(row=r, column=c1)
                l.border = Border(left=thin, top=l.border.top, right=l.border.right, bottom=l.border.bottom)
                rgt = ws.cell(row=r, column=c2)
                rgt.border = Border(right=thin, top=rgt.border.top, left=rgt.border.left, bottom=rgt.border.bottom)

        def box_all(r1, c1, r2, c2):
            for r in range(r1, r2+1):
                for c in range(c1, c2+1):
                    ws.cell(row=r, column=c).border = border_all

        # Mes en español
        MESES_ES = ["enero","febrero","marzo","abril","mayo","junio","julio","agosto",
                    "septiembre","octubre","noviembre","diciembre"]
        mes_es = MESES_ES[fecha.month-1]

        wb = Workbook()
        ws = wb.active; ws.title = "Lista"
        ws.sheet_view.showGridLines = False

        # Página
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_margins.left = ws.page_margins.right = 0.3
        ws.page_margins.top = ws.page_margins.bottom = 0.4

        # Columnas
        widths = {"A": 2, "B": 6, "C": 22, "D": 22, "E": 22, "F": 18, "G": 22,
                  "H": 20, "I": 16, "J": 6, "K": 6, "L": 10, "M": 6, "N": 6, "O": 6,
                  "P": 14, "Q": 14, "R": 14, "S": 16}
        for col, w in widths.items(): ws.column_dimensions[col].width = w

        # Alturas (logos un poquito más grandes)
        ws.row_dimensions[1].height = 8
        ws.row_dimensions[3].height = 50  # ↑ más aire para logos
        ws.row_dimensions[4].height = 22
        ws.row_dimensions[5].height = 18
        ws.row_dimensions[6].height = 14

        # ------- Logos y títulos centrados (B3..S5) -------
        try:
            if _Path("logo_izq.png").exists():
                img = XLImage("logo_izq.png")
                target_h = 72  # ↑ un poco más grande
                ratio = target_h / img.height
                img.height = target_h
                img.width  = int(img.width * ratio)
                ws.add_image(img, "D3")

            if _Path("logo_der.png").exists():
                img2 = XLImage("logo_der.png")
                target_h2 = 72  # ↑ un poco más grande
                ratio2 = target_h2 / img2.height
                img2.height = target_h2
                img2.width  = int(img2.width * ratio2)
                ws.add_image(img2, "O3")
        except Exception:
            pass

        # Títulos
        ws.merge_cells("B3:S3"); ws["B3"].value = "Modelo de Gestión Policial de Fuerza Pública"; ws["B3"].alignment=center; ws["B3"].font=h1_font
        ws.merge_cells("B4:S4"); ws["B4"].value = "Lista de Asistencia & Minuta"; ws["B4"].alignment=center; ws["B4"].font=h1_font
        ws.merge_cells("B5:S5"); ws["B5"].value = "Consecutivo:"; ws["B5"].alignment=center; ws["B5"].font=title_font

        # Banda azul
        ws.merge_cells("B6:S6"); ws["B6"].fill = PatternFill("solid", fgColor=azul_banda)

        # Marco del bloque superior (borde superior en FILA 1)
        outline_box(1, 2, 6, 19)

        # ======= Encabezado con cuadrícula =======
        ws.merge_cells(start_row=7, start_column=2, end_row=7, end_column=4)   # B7:D7
        ws.merge_cells(start_row=7, start_column=5, end_row=7, end_column=9)   # E7:I7
        ws.merge_cells(start_row=7, start_column=10, end_row=7, end_column=15) # J7:O7
        ws.merge_cells(start_row=7, start_column=16, end_row=7, end_column=19) # P7:S7
        ws["B7"].value = f"Fecha: {fecha.day} {mes_es} {fecha.year}"; ws["B7"].font = title_font; ws["B7"].alignment = left
        ws["E7"].value = f"Lugar:  {lugar}" if lugar else "Lugar: "; ws["E7"].font = title_font; ws["E7"].alignment = left
        ws["J7"].value = f"Hora Inicio: {hora_ini.strftime('%H:%M')}"; ws["J7"].alignment = center
        ws["P7"].value = f"Hora Finalización: {hora_fin.strftime('%H:%M')}"; ws["P7"].alignment = center
        box_all(7, 2, 7, 4); box_all(7, 5, 7, 9); box_all(7, 10, 7, 15); box_all(7, 16, 7, 19)

        # Fila 8: Estrategia
        ws.merge_cells(start_row=8, start_column=2, end_row=8, end_column=3)
        ws.merge_cells(start_row=8, start_column=4, end_row=8, end_column=9)
        ws["B8"].value = "Estrategia o Programa:"; ws["B8"].alignment = left
        ws["D8"].value = estrategia; ws["D8"].alignment = left
        box_all(8, 2, 8, 3); box_all(8, 4, 8, 9)

        # Actividad (solo contorno)
        ws.merge_cells(start_row=8, start_column=10, end_row=9, end_column=19)
        ws["J8"].value = "ACTIVIDAD: Reunión Virtual de Seguimiento de líneas de acción, acciones estratégicas, indicadores y metas."
        ws["J8"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        outline_box(8, 10, 9, 19)

        # Fila 9: Delegación
        ws.merge_cells(start_row=9, start_column=2, end_row=9, end_column=3)
        ws.merge_cells(start_row=9, start_column=4, end_row=9, end_column=9)
        ws["B9"].value = "Dirección / Delegación Policial:"; ws["B9"].alignment = left
        ws["D9"].value = delegacion; ws["D9"].alignment = Alignment(horizontal="center")
        box_all(9, 2, 9, 3); box_all(9, 4, 9, 9)

        # Encabezado de la tabla
        ws["B10"].value = ""
        ws["B10"].alignment = right
        ws.merge_cells("C10:E11"); ws["C10"].value = "Nombre"

        ws["F10"].value = "Cédula de Identidad"
        ws["G10"].value = "Institución"
        ws["H10"].value = "Cargo"
        ws["I10"].value = "Teléfono"
        ws.merge_cells("J10:L10"); ws["J10"].value = "Género"
        ws.merge_cells("M10:O10"); ws["M10"].value = "Sexo (Hombre, Mujer o Intersex)"
        ws.merge_cells("P10:R10"); ws["P10"].value = "Rango de Edad"
        ws["S10"].value = "FIRMA"

        for rng in ["C10:E11","J10:L10","M10:O10","P10:R10"]:
            c = ws[rng.split(":")[0]]; c.font = th_font; c.alignment = center; c.fill = celda_fill
        for cell in ["F10","G10","H10","I10","S10"]:
            ws[cell].font = th_font; ws[cell].alignment = center; ws[cell].fill = celda_fill

        ws["J11"], ws["K11"], ws["L11"] = "F", "M", "LGBTIQ+"
        ws["M11"], ws["N11"], ws["O11"] = "H", "M", "I"
        ws["P11"], ws["Q11"], ws["R11"] = "18 a 35 años", "36 a 64 años", "65 años o más"
        for cell in ["J11","K11","L11","M11","N11","O11","P11","Q11","R11"]:
            ws[cell].font = th_font; ws[cell].alignment = center; ws[cell].fill = celda_fill

        for r in range(10, 12):
            for c in range(2, 20):
                ws.cell(row=r, column=c).border = border_all

        # Congelar encabezado
        ws.freeze_panes = "A12"

        # Filas de asistencia
        start_row = 12
        for i, (_, row) in enumerate(rows_df.iterrows()):
            r = start_row + i
            ws[f"B{r}"].value = i + 1
            ws[f"B{r}"].alignment = right

            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
            ws[f"C{r}"].value = str(row.get("Nombre",""))
            ws[f"C{r}"].alignment = Alignment(wrap_text=True, horizontal="left", vertical="top")

            ws[f"F{r}"].value = str(row.get("Cédula de Identidad",""))
            ws[f"G{r}"].value = str(row.get("Institución",""))
            ws[f"H{r}"].value = str(row.get("Cargo",""))
            ws[f"I{r}"].value = str(row.get("Teléfono",""))

            for col in ["J","K","L","M","N","O","P","Q","R"]:
                ws[f"{col}{r}"].value = ""

            g = (row.get("Género","") or "").strip()
            if g == "F": ws[f"J{r}"].value = "X"
            elif g == "M": ws[f"K{r}"].value = "X"
            elif g == "LGBTIQ+": ws[f"L{r}"].value = "X"

            s = (row.get("Sexo","") or "").strip()
            if s == "H": ws[f"M{r}"].value = "X"
            elif s == "M": ws[f"N{r}"].value = "X"
            elif s == "I": ws[f"O{r}"].value = "X"

            e = (row.get("Rango de Edad","") or "").strip()
            if e.startswith("18"): ws[f"P{r}"].value = "X"
            elif e.startswith("36"): ws[f"Q{r}"].value = "X"
            elif e.startswith("65"): ws[f"R{r}"].value = "X"

            ws[f"S{r}"].value = "Virtual"

            for c in range(2, 20):
                ws.cell(row=r, column=c).border = border_all

        # ===== Anotaciones / Acuerdos (más bajos) =====
        last_data_row = start_row + len(rows_df) - 1 if len(rows_df) > 0 else 11
        notes_top = max(25, last_data_row + 2)
        notes_height = 14  # ← antes 20; ahora menos alto

        # Títulos de los cuadros
        ws.merge_cells(start_row=notes_top, start_column=2, end_row=notes_top, end_column=10)  # B..J
        ws.merge_cells(start_row=notes_top, start_column=12, end_row=notes_top, end_column=19) # L..S
        ws[f"B{notes_top}"].value = "Anotaciones Generales."; ws[f"B{notes_top}"].alignment = center
        ws[f"L{notes_top}"].value = "Acuerdos."; ws[f"L{notes_top}"].alignment = center
        ws[f"B{notes_top}"].font = th_font; ws[f"L{notes_top}"].font = th_font
        ws[f"B{notes_top}"].fill = celda_fill; ws[f"L{notes_top}"].fill = celda_fill

        # Marcos exteriores de los cuadros (sin cuadriculas internas)
        outline_box(notes_top+1, 2, notes_top+notes_height, 10)   # cuadro izq
        outline_box(notes_top+1, 12, notes_top+notes_height, 19)  # cuadro der

        # Rellenos de texto
        ws.merge_cells(start_row=notes_top+1, start_column=2, end_row=notes_top+notes_height, end_column=10)
        ws[f"B{notes_top+1}"].alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
        if anotaciones_txt.strip(): ws[f"B{notes_top+1}"].value = anotaciones_txt.strip()

        ws.merge_cells(start_row=notes_top+1, start_column=12, end_row=notes_top+notes_height, end_column=19)
        ws[f"L{notes_top+1}"].alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
        if acuerdos_txt.strip(): ws[f"L{notes_top+1}"].value = acuerdos_txt.strip()

        # ===== Pie (ajustado a la nueva altura de los cuadros) =====
        row_pie = notes_top + notes_height + 2
        for r in range(row_pie, row_pie + 8):
            for c in range(2, 20):
                ws.cell(row=r, column=c).border = Border()  # sin bordes

        ws.merge_cells(start_row=row_pie, start_column=2, end_row=row_pie, end_column=10)
        ws[f"B{row_pie}"].value = f"Se Finaliza la Reunión a:   {hora_fin.strftime('%H:%M')}"
        ws[f"B{row_pie}"].alignment = left

        from openpyxl.utils import get_column_letter
        row_firma = row_pie + 3
        thin_line = Side(style="thin", color="000000")
        sig_c1, sig_c2 = 4, 10  # D..J
        ws.merge_cells(start_row=row_firma, start_column=sig_c1, end_row=row_firma, end_column=sig_c2)
        for c in range(sig_c1, sig_c2 + 1):
            ws.cell(row=row_firma, column=c).border = Border(bottom=thin_line)

        ws.merge_cells(start_row=row_firma+1, start_column=sig_c1, end_row=row_firma+1, end_column=sig_c2)
        ws[f"{get_column_letter(sig_c1)}{row_firma+1}"].value = "Nombre Completo y Firma"
        ws[f"{get_column_letter(sig_c1)}{row_firma+1}"].alignment = Alignment(horizontal="center", wrap_text=False)

        ws.merge_cells(start_row=row_firma+3, start_column=2, end_row=row_firma+3, end_column=10)
        ws[f"B{row_firma+3}"].value = "Cargo:"
        ws[f"B{row_firma+3}"].alignment = left

        ws.merge_cells(start_row=row_firma+5, start_column=12, end_row=row_firma+5, end_column=19)
        ws[f"L{row_firma+5}"].value = "Sello Policial"
        ws[f"L{row_firma+5}"].alignment = Alignment(horizontal="right")

        # Protección
        ws.protection.sheet = True
        ws.protection.formatColumns = False
        ws.protection.formatRows = False
        ws.protection.selectLockedCells = True
        ws.protection.selectUnlockedCells = True

        bio = BytesIO(); wb.save(bio); return bio.getvalue()

    df_for_excel = fetch_all_df(include_id=False)
    datos = df_for_excel.drop(columns=["Nº"]) if not df_for_excel.empty else df_for_excel
    if st.button("📥 Generar y descargar Excel oficial", use_container_width=True, type="primary"):
        xls_bytes = build_excel_oficial_single(
            fecha_evento, lugar, hora_inicio, hora_fin, estrategia, delegacion, datos,
            anotaciones, acuerdos
        )
        if xls_bytes:
            st.download_button(
                "⬇️ Descargar Excel (una sola hoja)",
                data=xls_bytes,
                file_name=f"Lista_Asistencia_Oficial_{date.today():%Y%m%d}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )








